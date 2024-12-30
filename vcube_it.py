import pandas as pd
import os
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO
from html2image import Html2Image

# Ensure the directory for images exists
image_dir = "extracted_images"
os.makedirs(image_dir, exist_ok=True)

# Load the Excel workbook
workbook_path = 'C:/Users/manik/Downloads/Book1.xlsx'
wb = load_workbook(workbook_path)
sheet = wb.active  # Access the active sheet

# Extract images and map them to student names
image_map = {}
for image in sheet._images:  # Access embedded images
    row = image.anchor._from.row + 1  # Get the row where the image is located (1-based index)
    binary_data = image._data()  # Access raw binary data of the image

    # Convert binary data to an actual image
    try:
        img = Image.open(BytesIO(binary_data))
        # Get the student name from the sheet (assuming 'Name' is in the first column)
        student_name = sheet.cell(row=row, column=1).value
        if student_name:
            image_name = f"{student_name}.png"  # Use the student's name for the image filename
            image_path = os.path.join(image_dir, image_name)
            img.save(image_path)  # Save the image to a file
            image_map[row] = image_path  # Map the image path to the row
        else:
            image_map[row] = 'default_image.png'  # Default image for missing student name
    except Exception as e:
        print(f"Failed to process image for row {row}: {e}")
        image_map[row] = 'default_image.png'  # Assign a default image path on failure

# Load the Excel data as a DataFrame
data = pd.read_excel(workbook_path)

# Initialize Html2Image without specifying temporary_path
hti = Html2Image(output_path="id_cards_images")  # Set the output directory for images

# Start creating the HTML page content
html_page = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.0.2/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-EVSTQN3/azprG1Anm3QDgpJLIm9Nao0Yz1ztcQTwFspd3yD65VohhpuuCOmLASjC" crossorigin="anonymous">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Document</title>
    <style>
        .con{
            width: 300px;
            height: 460px;
            border: 1px solid black;
            border-radius: 2px;
            background-color: rgb(254, 177, 35);
        }
        .top-logo{
            width: 100%;
            height: 160px;
            background-color:whitesmoke;
        }
        .con .top-logo .std-img-div{
            height: 150px;
            width: 150px;
            border: 5px solid orange;
            background-color: aliceblue;
            border-radius: 15px;
            
            position: absolute;
            top: 150px;
            left:76px;
        }
        .std-img-div img {
            height: 140px; /* Ensure the image fills the container */
            width: 141px;  /* Maintain responsiveness */
            object-fit: cover; /* Crop the image to fit if needed */
            border-radius: 10px; /* Same as the container for a clean look */
        }
        .vcube-logo{
            height: 120px;
            width: 270px;
        }
    </style>
</head>
<body>
"""

# Loop through each student data and create an ID card
for index, row in data.iterrows():
    # Extract student information, providing 'None' for missing data
    student_name = row.get('Name', 'None')  # Default to 'None' if missing
    student_id = row.get('ID', 'None')  # Default to 'None' if missing
    student_role = row.get('Role', 'None')  # Default to 'None' if missing
    row_number = index + 2  # Adjusted to match the row numbers in the image map (Excel row starts at 2)
    
    # Get the correct image path for the student from the image map
    student_image = image_map.get(row_number, 'default_image.png')  # Use extracted image or default

    # Generate the HTML content for each student ID card
    card_html = f"""
    <div class="con" id="card-{index}">
        <div class="top-logo">
            <img class="vcube-logo" src="Screenshot 2024-12-27 095107.png" alt="">
            <div class="std-img-div">
                <img src="{student_image}" alt="">
            </div>
        </div>
        <div class="bottom-content" style="padding: 20px; margin-top: 130px;">
            <h6 style="margin-left:40px;">Name: {student_name} </h6>
            <h6 style="margin-left:40px;">Role:{student_role} </h6>
            <h6 style="margin-left:40px;">Employeeid: {student_id}</h6>
        </div>
    </div>
    <div>        <button class="btn btn-primary" id="downloadBtn-{index}" onclick="downloadCard({index})">Download Card</button>
    </div>
    """

    # Add the card HTML to the main page
    html_page += card_html

# Add the JavaScript for downloading cards
html_page += """
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<script>
    function downloadCard(index) {
        const card = document.getElementById(`card-${index}`); // Target the specific card

        // Use html2canvas to capture the entire card element
        html2canvas(card, {
            useCORS: true, // Ensure cross-origin images are included
            scale: 2 // Improve the quality of the generated image
        }).then(canvas => {
            const link = document.createElement('a');
            link.download = 'Vcube_ID-Card.png'; // Set the download filename
            link.href = canvas.toDataURL('image/png'); // Convert canvas to image
            link.click(); // Trigger the download
        });
    }
</script>
</body>
</html>
"""

# Write the HTML content to an output file
with open('output.html', 'w') as file:
    file.write(html_page)

print("Student ID cards HTML has been generated and saved to 'output.html'!")
